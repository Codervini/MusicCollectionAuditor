from pathlib import Path
import logging
import re
import csv
from dotenv import dotenv_values
from collections import defaultdict
import shutil
import datetime
import openpyxl
import send2trash
CONFIG_CONSTANTS = dotenv_values(".env")

def csvFormatter(name):
    return f'logs_test/{name} {datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.csv'

def writeAllToExcel(song_full_path,fail_record_lookup):
    if CONFIG_CONSTANTS["QCLEVEL_3_CENTRAL_EXCEL"]:
        excel_file = CONFIG_CONSTANTS["QCLEVEL_3_CENTRAL_EXCEL"]
    else:
        excel_file =  Path("data","song list.xlsx")

    if Path(excel_file).exists():
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        excel_file_value_lookup = { tuple(i[0:2]): i[2] for i in sheet.iter_rows(2,values_only=True)}
        for i in song_full_path:
            if (str(Path(*i[:-1])), i[-1]) not in excel_file_value_lookup and fail_record_lookup.get(tuple(i), "Clean") != excel_file_value_lookup.get((str(Path(*i[:-1])), i[-1])):
                sheet.append([str(Path(*i[:-1])), i[-1],fail_record_lookup.get(tuple(i), "Clean")])
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["File Path", "Song", "Result"])
        for i in song_full_path:
            sheet.append([str(Path(*i[:-1])), i[-1],fail_record_lookup.get(tuple(i), "Clean")])
            # print((str(Path(*i[:-1])), i[-1])) 
    try:
        wb.save(excel_file)
        return True
    except:
        return False


def writeAllToCSV(song_full_path, fail_record_lookup):
    csv_file_value_lookup = {}
    is_header_available = False
    if CONFIG_CONSTANTS["QCLEVEL_3_CENTRAL_CSV"]:
        csv_file = CONFIG_CONSTANTS["QCLEVEL_3_CENTRAL_CSV"]
        try:
            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                reader_list = list(reader)
                if len(reader_list) > 0:
                    if ["File Path", "Song", "Result"] ==  reader_list[0]:
                        is_header_available = True                                        
                csv_file_value_lookup = { tuple(i[:-1]): i[-1] for i in reader_list[1:]}
        except FileNotFoundError:
            pass

        with open(csv_file,"a") as f:
            writer = csv.writer(f,delimiter=",")
            if not is_header_available:
                writer.writerow(["File Path", "Song", "Result"])
            for i in song_full_path:
                if (str(Path(*i[:-1])), i[-1]) not in csv_file_value_lookup and  fail_record_lookup.get(tuple(i), "Clean") != csv_file_value_lookup.get((str(Path(*i[:-1])), i[-1])):
                    writer.writerow([str(Path(*i[:-1])), i[-1],fail_record_lookup.get(tuple(i), "Clean")])
        return True
    else:
        csv_file = csvFormatter("all_songs")
        with open(csv_file,"a") as f:
            writer = csv.writer(f,delimiter=",")
            if not is_header_available:
                writer.writerow(["File Path", "Song", "Result"])
            for i in song_full_path:
                writer.writerow([str(Path(*i[:-1])), i[-1],fail_record_lookup.get(tuple(i), "Clean")])
        return True
            

def QualityChecker(FLAC_TEST_REPORT,CONFIG_CONSTANTS):
    # FLAC_TEST_REPORT = input("Enter the FLAC-authenticity CSV report absolute path: ")
    # top_root_folder = ""
    song_full_path =[]
    try:
        with open(Path(FLAC_TEST_REPORT).expanduser(),"r") as f:
            headerFormat = r"^Lossless Audio Checker \d\.\d\.\d logfile from [A-Za-z]+ \d{1,2} [A-Za-z]+ \d{4} \d{2}:\d{2}:\d{2} (AM|PM)\n$"
            logging.info("FLAC-authenticity CSV report accesed")
            records = f.readlines()
            records_count = 0
            result_record = {}
            fail_record = []

            for i in records:
                if not i.isspace():
                    if re.fullmatch(headerFormat,i):
                        logging.info("Log File provided. It is valid. Source: " + " ".join(i.split()[:4]))
                    elif i.split()[0] == "File:":
                        formatted_file_path = i.replace(" \\","").replace("\n","").split("\\")[1:]
                        records_count += 1
                        song_full_path.append(formatted_file_path)
                    elif i.split()[0] == "Result:":
                        formated_result = i.split()[-1]
                        if formated_result != "Clean":
                            fail_record.append([formatted_file_path,formated_result])

                        if formated_result not in result_record:
                            result_record[formated_result] = 1
                        elif formated_result in result_record:
                            result_record[formated_result] += 1
                        else:
                            logging.error("Something happened")
                    else:
                        logging.error("Unknown report provided. Please provide supported report.")
                        break
            else:
                logging.info("Processing of FLAC report completed succesfully")

                if len(song_full_path) > 1:
                    path_dicer_count = defaultdict(int)
                    for i in song_full_path:
                        for j in i:
                            path_dicer_count[j] += 1
                    keys = list(path_dicer_count.keys())
                    for i in zip(keys, keys[1:]):
                        if path_dicer_count[i[0]] != path_dicer_count[i[1]] and path_dicer_count[i[0]] == records_count:
                            top_common_folder = i[0]
                else:
                    top_common_folder = list(path_dicer_count.keys())[-2]

                top_common_folder_index = song_full_path[0].index(top_common_folder)
                path_to_common_folder = song_full_path[0][0: top_common_folder_index + 1]
                if CONFIG_CONSTANTS["QC_ACTUAL_ABSOLUTE_PATH_TO_TOP_ROOT_FOLDER"]:
                    temp_ptcf = list(Path(CONFIG_CONSTANTS["QC_ACTUAL_ABSOLUTE_PATH_TO_TOP_ROOT_FOLDER"]).parts) 
                    path_to_common_folder = [*temp_ptcf, *path_to_common_folder[path_to_common_folder.index(top_common_folder)+1:]]
                    for i in range(0, len(song_full_path)):
                        song_full_path[i] = [*path_to_common_folder, *song_full_path[i][top_common_folder_index + 1:]]
                    for i in fail_record:
                        i[0] = [*path_to_common_folder, *i[0][top_common_folder_index + 1:]]
                    logging.info("Alternative absolute path provided. Changed the records to reflect the same.")
                logging.info(f"Path to Common folder: {path_to_common_folder}")



                # print(type(path_to_common_folder))
                for i in fail_record:
                    # print(f"{"\\".join(i[0])}: {i[1]}")
                    print(f"{str(Path(*i[0]))}: {i[1]}")
                    logging.info(f"{str(Path(*i[0]))}: {i[1]}")
                # print("")
                print(f"Total {records_count} songs found of which ",end="")
                logging.info(f"Total {records_count} songs found of which ")
                for i in result_record:
                    print(f"[ {result_record[i]} {i} ] ", end="")
                    logging.info(f"[ {result_record[i]} {i} ] ")
                print("\n")

                fail_record_lookup={tuple(i[0]): i[1] for i in fail_record}

                logging.info("FLAC report findings provided.")

                match int(CONFIG_CONSTANTS["QCLEVEL_2"]):
                    case 4 | 5 | 6:
                        qcLevel_2 = CONFIG_CONSTANTS["QCLEVEL_2"]
                    case _:
                        qcLevel_2 = int(input("""What do you want to do next?
1 - to proceed
2 - to provide a csv of the failed songs
3 - to provide a csv of the failed songs and delete the failed songs in the folder
4 - to provide a csv of all of the songs (including failed)
5 - to provide an excel file of all the songs (including failed)
6 - to provide a csv and excel sheet of all the songs (including failed)
Your choice: """).strip())
                
                match qcLevel_2:
                    case 1:
                        pass
                        logging.info("User choose to proceed")
                    case 2:
                        with open(csvFormatter("failed_songs2"),"w",newline="") as f:
                            writer =  csv.writer(f,delimiter=",")
                            writer.writerow(["File Path", "Song", "Result"])
                            for i in fail_record:
                                # writer.writerow(["\\".join(i.split("\\")[:-1]), i.split("\\")[-1], fail_record[i]])
                                # writer.writerow(["\\".join(i[0][:-1]),i[0][-1],i[1]])
                                writer.writerow([str(Path(*i[0][:-1])),i[0][-1],i[1]])  
                        logging.info("User choose to receive a csv of the failed songs")    

                    case 3:
                        with open(csvFormatter("failed_songs3"),"w",newline="") as f:
                            writer =  csv.writer(f,delimiter=",")
                            writer.writerow(["File Path", "Song", "Result"])
                            for i in fail_record:
                                # writer.writerow(["/".join(i[0][:-1]),i[0][-1],i[1]])
                                writer.writerow([str(Path(*i[0][:-1])),i[0][-1],i[1]])
                        
                        deleted_file_count = 0
                        if bool(CONFIG_CONSTANTS["QCLEVEL_3_MOVE_TO__LOCAL_TRASH_FOLDER_DELETE"]):
                            Path(*path_to_common_folder, ".deletedSongs").expanduser().mkdir(exist_ok=True)
                            if len(fail_record) > 0:
                                for i in fail_record:
                                    try:
                                        shutil.move(Path(*i[0]).expanduser(),Path(*path_to_common_folder, ".deletedSongs").expanduser())
                                        deleted_file_count += 1
                                    except FileNotFoundError as e:
                                        # print(f"Warning: {Path("/".join(i[0])).expanduser()} not found. Skipping.")
                                        logging.warning(f"{Path(*i[0]).expanduser()} not found. Skipping.{e}")
                                    except shutil.Error as e:
                                        logging.warning(f"{Path(*i[0]).expanduser()} already moved in. Skipping. {e}")
                                else:
                                    print(f"Deleted {deleted_file_count} files to .deletedSongs")
                                    logging.info(f"Deleted {deleted_file_count} files to .deletedSongs")
                            else:
                                print("No failed songs in report to delete")
                                logging.info("No failed songs in report to delete")
                        else:
                            for i in fail_record:
                                try:
                                    send2trash(Path(*i[0]).expanduser())
                                    deleted_file_count += 1
                                except FileNotFoundError as e:
                                    logging.warning(f"{Path(*i[0]).expanduser()} not found. Skipping.{e}")
                        logging.info("User choose to receive a csv of the failed songs and delete the failed songs in the folder")
                    case 4:
                        if writeAllToCSV(song_full_path,fail_record_lookup):
                            logging.info("User choose to receive a csv of all of the songs (including failed)")
                        else:
                            logging.error("Something went wrong with CSV process")

                    case 5:    
                        if writeAllToExcel(song_full_path,fail_record_lookup):
                            logging.info("User choose to receive an excel file of all the songs (including failed)")
                        else:
                            logging.error("Something went wrong with excel process")

                    case 6:
                        logging.info("User choose to to receive a csv and excel sheet of all the songs (including failed)")

                        if writeAllToExcel(song_full_path,fail_record_lookup):
                            logging.info("Excel provided 1/2")
                        else:
                            logging.error("Something went wrong with excel process 1/2")

                        if writeAllToCSV(song_full_path,fail_record_lookup):
                            logging.info("CSV provided 2/2")
                        else:
                            logging.error("Something went wrong with CSV process 2/2")


                    case _:
                        logging.info("User choose invalid option")
    except FileNotFoundError as e:
        logging.error(f"FLAC-authenticity CSV report not given {e}")

# QualityChecker("Lossless Audio Checker.log",CONFIG_CONSTANTS)